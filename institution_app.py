#!/usr/bin/env python3
"""
Institution RCR Analyzer - Streamlit App
=========================================
Enter institution names to calculate institution-level iCite / RCR metrics
for all publications from the last 2 full calendar years.

Run with:
    streamlit run institution_app.py

Requirements:
    pip install streamlit openpyxl requests
"""

import io
import re
import time
import statistics
import xml.etree.ElementTree as ET
from datetime import datetime

import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------
ESEARCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
EFETCH_URL = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/efetch.fcgi"
ICITE_URL = "https://icite.od.nih.gov/api/pubs"
ICITE_BATCH = 200

CURRENT_YEAR = datetime.now().year
ANALYSIS_YEARS = [CURRENT_YEAR - 2, CURRENT_YEAR - 1]  # last 2 full calendar years


# ---------------------------------------------------------------------------
# Institution name variant generator
# ---------------------------------------------------------------------------
_STATE_ABBREVS = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
    "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
    "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
    "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
    "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
    "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
    "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
    "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
    "District of Columbia": "DC",
}

_SKIP_WORDS = {"of", "the", "and", "at", "in", "for", "on", "&"}

_KNOWN_UC_ABBREVS = {
    "San Francisco": "UCSF", "Los Angeles": "UCLA", "San Diego": "UCSD",
    "Santa Barbara": "UCSB", "Santa Cruz": "UCSC", "Davis": "UCD",
    "Irvine": "UCI", "Riverside": "UCR", "Berkeley": "UCB", "Merced": "UCM",
}

# Well-known institution aliases not easily derived by rules
_KNOWN_ALIASES = {
    "university of pennsylvania": ["UPenn", "Penn Medicine", "Penn Med",
                                    "Perelman School of Medicine"],
    "university of michigan": ["Michigan Medicine", "U-M", "UMich"],
    "yale university": ["Yale-New Haven Hospital", "Yale New Haven",
                        "Yale Medicine"],
    "stanford university": ["Stanford Health Care", "Stanford Medicine"],
    "columbia university": ["Columbia University Irving Medical Center",
                            "NewYork-Presbyterian", "Columbia Irving"],
    "harvard university": ["Harvard Medical School", "Harvard T.H. Chan",
                           "Harvard T H Chan", "Beth Israel Deaconess",
                           "Brigham and Women", "Massachusetts General",
                           "Boston Children's Hospital", "Dana-Farber",
                           "McLean Hospital"],
    "johns hopkins university": ["Johns Hopkins Hospital",
                                  "Johns Hopkins Medicine",
                                  "Johns Hopkins Bloomberg"],
    "duke university": ["Duke Health", "Duke University Medical Center",
                        "Duke University Hospital"],
    "university of pittsburgh": ["UPMC", "Pitt"],
    "cornell university": ["Weill Cornell", "NewYork-Presbyterian/Weill Cornell"],
    "university of washington": ["UW Medicine", "UW Medical Center",
                                  "Harborview Medical Center"],
    "mayo clinic": ["Mayo Clinic Rochester", "Mayo Clinic Arizona",
                    "Mayo Clinic Florida", "Mayo Foundation"],
    "massachusetts general hospital": ["Mass General", "Mass General Brigham",
                                        "MGH"],
    "mount sinai": ["Icahn School of Medicine at Mount Sinai",
                    "Icahn School of Medicine", "Mount Sinai Hospital",
                    "Mount Sinai Health System"],
    "university of california, los angeles": ["Ronald Reagan UCLA Medical Center"],
    "university of california, san diego": ["UC San Diego Health"],
    "university of california, san francisco": ["UCSF Benioff", "San Francisco General Hospital",
                                                  "Zuckerberg San Francisco General"],
    "emory university": ["Emory Healthcare", "Emory School of Medicine",
                         "Grady Memorial Hospital"],
    "university of chicago": ["UChicago Medicine", "UChicago"],
    "northwestern university": ["Northwestern Medicine",
                                 "Northwestern Memorial Hospital",
                                 "Feinberg School of Medicine"],
    "vanderbilt university": ["Vanderbilt University Medical Center", "VUMC",
                               "Vanderbilt Health"],
    "new york university": ["NYU Langone", "NYU Grossman School of Medicine",
                             "NYU Grossman"],
    "university of texas southwestern": ["UT Southwestern", "UTSW",
                                          "Parkland Memorial Hospital"],
    "university of colorado": ["UCHealth", "UC Denver",
                                "University of Colorado Anschutz"],
    "baylor college of medicine": ["Baylor College", "Texas Children's Hospital"],
    "washington university in st. louis": ["Washington University in St Louis",
                                            "WashU", "Barnes-Jewish Hospital",
                                            "BJC HealthCare"],
    "university of virginia": ["UVA Health", "UVA"],
    "university of iowa": ["UI Health Care", "University of Iowa Hospitals"],
    "university of minnesota": ["M Health Fairview", "UMN"],
    "university of wisconsin": ["UW Health", "UW-Madison"],
    "oregon health & science university": ["OHSU"],
    "oregon health and science university": ["OHSU"],
    "university of alabama at birmingham": ["UAB", "UAB Medicine"],
    "university of rochester": ["UR Medicine", "Strong Memorial Hospital"],
    "university of florida": ["UF Health", "UF College of Medicine"],
    "university of maryland": ["University of Maryland Medical Center", "UMMC"],
    "case western reserve university": ["CWRU", "University Hospitals Cleveland",
                                         "MetroHealth"],
    "brown university": ["Warren Alpert Medical School", "Lifespan",
                         "Rhode Island Hospital", "Brown Medicine"],
    "university of cincinnati": ["UC Health", "UC College of Medicine"],
    "tufts university": ["Tufts Medical Center", "Tufts School of Medicine"],
    "georgetown university": ["MedStar Georgetown", "Georgetown University Medical Center"],
    "boston university": ["Boston Medical Center", "BU School of Medicine",
                          "Boston University Chobanian"],
    "thomas jefferson university": ["Jefferson Health", "Sidney Kimmel Medical College"],
    "rush university": ["Rush University Medical Center", "Rush Medical College"],
    "tulane university": ["Tulane School of Medicine", "Tulane Medical Center"],
    "university of south florida": ["USF Health", "USF Morsani College of Medicine"],
    "university of utah": ["U of U Health", "University of Utah Health",
                            "Huntsman Cancer Institute"],
    "wake forest university": ["Atrium Health Wake Forest Baptist",
                                "Wake Forest School of Medicine",
                                "Wake Forest Baptist"],
    "medical university of south carolina": ["MUSC", "MUSC Health"],
    "university of nebraska": ["UNMC", "Nebraska Medicine",
                                "University of Nebraska Medical Center"],
    "university of kansas": ["KU Medical Center", "University of Kansas Health System"],
    "university of kentucky": ["UK HealthCare", "UK College of Medicine"],
    "university of arkansas": ["UAMS", "University of Arkansas for Medical Sciences"],
    "ohio state university": ["Ohio State Wexner Medical Center", "OSU"],
    "pennsylvania state university": ["Penn State Health", "Penn State Hershey",
                                       "Penn State College of Medicine"],
    "michigan state university": ["MSU College of Human Medicine",
                                   "Sparrow Hospital"],
    "indiana university": ["IU School of Medicine", "IU Health"],
    "university of tennessee": ["UT Health Science Center", "UTHSC"],
    "university of texas": ["UT Health", "Dell Medical School"],
}


def generate_institution_variants(name):
    """Generate common affiliation variants for an institution name."""
    name = name.strip()
    if not name:
        return []

    variants = set()
    variants.add(name)

    # Add known aliases
    name_lower = name.lower().rstrip(".")
    for key, aliases in _KNOWN_ALIASES.items():
        if key in name_lower or name_lower in key:
            for alias in aliases:
                variants.add(alias)

    # Remove commas for a no-comma variant
    no_comma = name.replace(",", "")
    if no_comma != name:
        variants.add(no_comma)

    # --- "University of [State/Place], [City]" pattern ---
    uc_match = re.match(
        r"University of ([A-Za-z ]+?)(?:,\s*(.+))?$", name, re.IGNORECASE
    )
    if uc_match:
        state_or_place = uc_match.group(1).strip()
        city = (uc_match.group(2) or "").strip()

        if city:
            variants.add(f"University of {state_or_place} {city}")
            if city in _KNOWN_UC_ABBREVS:
                abbrev = _KNOWN_UC_ABBREVS[city]
                variants.add(abbrev)
                variants.add(f"{abbrev} Medical Center")
                variants.add(f"{abbrev} Health")
                variants.add(f"{abbrev} School of Medicine")
            else:
                state_abbrev = "".join(
                    w[0].upper() for w in state_or_place.split() if w.lower() not in _SKIP_WORDS
                )
                city_abbrev = "".join(
                    w[0].upper() for w in city.split() if w.lower() not in _SKIP_WORDS
                )
                uc_short = f"U{state_abbrev}"
                variants.add(uc_short)
                variants.add(f"{uc_short}{city_abbrev}")
                variants.add(f"{uc_short} {city}")
            if state_or_place.lower() == "california":
                variants.add(f"UC {city}")
        else:
            state_abbrev = "".join(
                w[0].upper() for w in state_or_place.split() if w.lower() not in _SKIP_WORDS
            )
            variants.add(f"U{state_abbrev}")
            if state_or_place in _STATE_ABBREVS:
                st_code = _STATE_ABBREVS[state_or_place]
                variants.add(f"U{st_code}")

        if city:
            variants.add(f"University of {state_or_place} {city} School of Medicine")
            variants.add(f"University of {state_or_place}, {city} School of Medicine")
        else:
            variants.add(f"University of {state_or_place} School of Medicine")

    # --- "[Place] University" pattern ---
    place_univ_match = re.match(r"(.+?)\s+University$", name, re.IGNORECASE)
    if place_univ_match and not uc_match:
        place = place_univ_match.group(1).strip()
        variants.add(place)
        variants.add(f"{place} University School of Medicine")
        variants.add(f"{place} School of Medicine")
        variants.add(f"{place} Medical School")
        variants.add(f"{place} Medical Center")
        variants.add(f"{place} University Medical Center")
        words = [w for w in place.split() if w.lower() not in _SKIP_WORDS]
        if len(words) > 1:
            abbrev = "".join(w[0].upper() for w in words)
            variants.add(abbrev)

    # --- "[Name] School of Medicine" pattern ---
    som_match = re.match(r"(.+?)\s+School of Medicine$", name, re.IGNORECASE)
    if som_match:
        base = som_match.group(1).strip()
        variants.add(base)
        variants.add(f"{base} College of Medicine")
        variants.add(f"{base} Medical School")
        variants.add(f"{base} Medical Center")

    # --- "[Name] Medical Center" / "[Name] Hospital" pattern ---
    mc_match = re.match(r"(.+?)\s+(Medical Center|Hospital|Health System|Health Sciences?)$", name, re.IGNORECASE)
    if mc_match:
        base = mc_match.group(1).strip()
        suffix = mc_match.group(2).strip()
        variants.add(base)
        if "Medical Center" in suffix:
            variants.add(f"{base} Hospital")
            variants.add(f"{base} Health")
        elif "Hospital" in suffix:
            variants.add(f"{base} Medical Center")
            variants.add(f"{base} Health")

    # --- Generic abbreviation ---
    words = name.replace(",", " ").split()
    significant = [w for w in words if w[0:1].isupper() and w.lower() not in _SKIP_WORDS]
    if len(significant) >= 2:
        abbrev = "".join(w[0].upper() for w in significant)
        if 2 <= len(abbrev) <= 6:
            variants.add(abbrev)

    variants.discard("")
    return sorted(variants)


# ---------------------------------------------------------------------------
# PubMed search with 10k workaround
# ---------------------------------------------------------------------------
def get_ncbi_delay(api_key):
    return 0.11 if api_key else 0.34


def esearch_count(query, email, api_key=None):
    params = {
        "db": "pubmed", "term": query, "rettype": "count",
        "retmode": "json", "email": email,
    }
    if api_key:
        params["api_key"] = api_key
    resp = requests.get(ESEARCH_URL, params=params, timeout=30)
    resp.raise_for_status()
    return int(resp.json().get("esearchresult", {}).get("count", 0))


def esearch_ids(query, email, api_key=None, retmax=10000):
    params = {
        "db": "pubmed", "term": query, "retmax": retmax,
        "retmode": "json", "email": email,
    }
    if api_key:
        params["api_key"] = api_key
    time.sleep(get_ncbi_delay(api_key))
    resp = requests.get(ESEARCH_URL, params=params, timeout=30)
    resp.raise_for_status()
    return resp.json().get("esearchresult", {}).get("idlist", [])


def build_affiliation_query(institution_names, year):
    affil_parts = []
    for name in institution_names:
        name = name.strip()
        if name:
            affil_parts.append(f'"{name}"[Affiliation]')
    affil_query = "(" + " OR ".join(affil_parts) + ")"
    date_query = f"{year}/01/01:{year}/12/31[pdat]"
    return f"{affil_query} AND {date_query}"


def build_monthly_query(institution_names, year, month):
    affil_parts = []
    for name in institution_names:
        name = name.strip()
        if name:
            affil_parts.append(f'"{name}"[Affiliation]')
    affil_query = "(" + " OR ".join(affil_parts) + ")"
    if month == 12:
        end_date = f"{year}/12/31"
    else:
        end_date = f"{year}/{month + 1:02d}/01"
    start_date = f"{year}/{month:02d}/01"
    date_query = f"{start_date}:{end_date}[pdat]"
    return f"{affil_query} AND {date_query}"


def search_institution_pmids(institution_names, year, email, api_key=None, status_fn=None):
    query = build_affiliation_query(institution_names, year)
    count = esearch_count(query, email, api_key)
    if status_fn:
        status_fn(f"  {year}: {count:,} publications found")
    if count == 0:
        return []
    if count <= 9500:
        return esearch_ids(query, email, api_key, retmax=10000)
    all_pmids = set()
    for month in range(1, 13):
        monthly_query = build_monthly_query(institution_names, year, month)
        monthly_count = esearch_count(monthly_query, email, api_key)
        if monthly_count == 0:
            continue
        if monthly_count > 9500 and status_fn:
            status_fn(f"  {year}/{month:02d}: {monthly_count:,} papers (exceeds 10k limit, results may be incomplete)")
        ids = esearch_ids(monthly_query, email, api_key, retmax=10000)
        all_pmids.update(ids)
        if status_fn:
            status_fn(f"  {year}/{month:02d}: retrieved {len(ids):,} PMIDs (total so far: {len(all_pmids):,})")
    return list(all_pmids)


# ---------------------------------------------------------------------------
# iCite
# ---------------------------------------------------------------------------
def fetch_icite(pmids, status_fn=None):
    result = {}
    total = len(pmids)
    for start in range(0, total, ICITE_BATCH):
        batch = pmids[start:start + ICITE_BATCH]
        params = {
            "pmids": ",".join(batch),
            "fl": "pmid,relative_citation_ratio,nih_percentile,citation_count,"
                  "expected_citations_per_year,citations_per_year,is_research_article,"
                  "year,title,journal,doi,is_clinical,provisional",
        }
        time.sleep(0.5)
        try:
            resp = requests.get(ICITE_URL, params=params, timeout=30)
            resp.raise_for_status()
            data = resp.json()
            for pub in data.get("data", []):
                result[str(pub.get("pmid", ""))] = pub
        except Exception:
            pass
        if status_fn and (start + ICITE_BATCH) % 1000 < ICITE_BATCH:
            status_fn(f"  iCite: processed {min(start + ICITE_BATCH, total):,} / {total:,} PMIDs")
    return result


# ---------------------------------------------------------------------------
# Metrics calculation
# ---------------------------------------------------------------------------
def compute_institution_metrics(icite_data, year=None):
    rcrs = []
    percentiles = []
    citation_counts = []
    research_count = 0
    clinical_count = 0
    provisional_count = 0

    for pmid, pub in icite_data.items():
        if year and pub.get("year") != year:
            continue
        rcr = pub.get("relative_citation_ratio")
        if rcr is not None:
            rcrs.append(rcr)
        pct = pub.get("nih_percentile")
        if pct is not None:
            percentiles.append(pct)
        cc = pub.get("citation_count")
        if cc is not None:
            citation_counts.append(cc)
        if pub.get("is_research_article"):
            research_count += 1
        if pub.get("is_clinical"):
            clinical_count += 1
        if pub.get("provisional"):
            provisional_count += 1

    total_pubs = len(icite_data) if year is None else sum(1 for p in icite_data.values() if p.get("year") == year)

    return {
        "total_publications": total_pubs,
        "pubs_with_rcr": len(rcrs),
        "rcr_sum": round(sum(rcrs), 2) if rcrs else 0,
        "rcr_mean": round(statistics.mean(rcrs), 4) if rcrs else 0,
        "rcr_median": round(statistics.median(rcrs), 4) if rcrs else 0,
        "percentile_mean": round(statistics.mean(percentiles), 2) if percentiles else 0,
        "percentile_median": round(statistics.median(percentiles), 2) if percentiles else 0,
        "citation_total": sum(citation_counts),
        "citation_mean": round(statistics.mean(citation_counts), 2) if citation_counts else 0,
        "research_articles": research_count,
        "clinical_articles": clinical_count,
        "provisional_rcr_count": provisional_count,
    }


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1B4F72")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
ALT_FILL = PatternFill("solid", fgColor="D6EAF8")


def write_institution_xlsx(institution_label, icite_data, metrics_by_year, combined_metrics):
    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"

    ws_sum.merge_cells("A1:F1")
    ws_sum["A1"] = f"{institution_label} \u2014 Publication Metrics"
    ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color="1B4F72")
    ws_sum["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws_sum.row_dimensions[1].height = 30

    ws_sum.merge_cells("A2:F2")
    ws_sum["A2"] = f"Analysis period: {ANALYSIS_YEARS[0]}\u2013{ANALYSIS_YEARS[-1]}"
    ws_sum["A2"].font = Font(name="Arial", italic=True, size=10, color="666666")

    headers = ["Metric"] + [str(y) for y in ANALYSIS_YEARS] + ["Combined"]
    col_widths = [30] + [18] * (len(ANALYSIS_YEARS) + 1)
    header_row = 4

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws_sum.cell(row=header_row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws_sum.column_dimensions[get_column_letter(col_idx)].width = width

    metric_rows = [
        ("Total Publications", "total_publications"),
        ("Publications with RCR", "pubs_with_rcr"),
        ("RCR Sum", "rcr_sum"),
        ("Mean RCR", "rcr_mean"),
        ("Median RCR", "rcr_median"),
        ("Mean NIH Percentile", "percentile_mean"),
        ("Median NIH Percentile", "percentile_median"),
        ("Total Citations", "citation_total"),
        ("Mean Citations per Paper", "citation_mean"),
        ("Research Articles", "research_articles"),
        ("Clinical Articles", "clinical_articles"),
        ("Provisional RCR Count", "provisional_rcr_count"),
    ]

    for i, (label, key) in enumerate(metric_rows):
        row = header_row + 1 + i
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        cell = ws_sum.cell(row=row, column=1, value=label)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.border = THIN_BORDER
        cell.fill = fill
        for j, year in enumerate(ANALYSIS_YEARS):
            val = metrics_by_year.get(year, {}).get(key, "N/A")
            cell = ws_sum.cell(row=row, column=2 + j, value=val)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            cell.fill = fill
        val = combined_metrics.get(key, "N/A")
        cell = ws_sum.cell(row=row, column=2 + len(ANALYSIS_YEARS), value=val)
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
        cell.fill = fill

    ws_pubs = wb.create_sheet("All Publications")
    pub_headers = ["PMID", "Year", "Title", "Journal", "DOI", "RCR", "NIH Percentile", "Citations", "Research", "Clinical"]
    pub_widths = [12, 8, 55, 25, 25, 10, 16, 12, 10, 10]

    for col_idx, (header, width) in enumerate(zip(pub_headers, pub_widths), start=1):
        cell = ws_pubs.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws_pubs.column_dimensions[get_column_letter(col_idx)].width = width

    sorted_pubs = sorted(icite_data.values(), key=lambda p: (p.get("year", 0), p.get("relative_citation_ratio") or 0), reverse=True)

    for i, pub in enumerate(sorted_pubs):
        row = 2 + i
        rcr = pub.get("relative_citation_ratio")
        pct = pub.get("nih_percentile")
        values = [
            int(pub.get("pmid", 0)),
            pub.get("year", ""),
            pub.get("title", ""),
            pub.get("journal", ""),
            pub.get("doi", ""),
            round(rcr, 2) if rcr is not None else "N/A",
            round(pct, 1) if pct is not None else "N/A",
            pub.get("citation_count", "N/A"),
            "Yes" if pub.get("is_research_article") else "No",
            "Yes" if pub.get("is_clinical") else "No",
        ]
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for col_idx, val in enumerate(values, start=1):
            cell = ws_pubs.cell(row=row, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.fill = fill
            if col_idx in (1, 2, 6, 7, 8, 9, 10):
                cell.alignment = Alignment(horizontal="center", vertical="top")
            else:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_pubs.freeze_panes = "A2"
    ws_pubs.auto_filter.ref = f"A1:J{1 + len(sorted_pubs)}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------
def run_institution_pipeline(institution_names, email, api_key, progress_bar, status_text):
    all_pmids = {}
    for i, year in enumerate(ANALYSIS_YEARS):
        status_text.text(f"Searching PubMed for {year} publications...")
        pmids = search_institution_pmids(
            institution_names, year, email, api_key,
            status_fn=lambda msg: status_text.text(msg),
        )
        all_pmids[year] = pmids
        progress_bar.progress((i + 1) / (len(ANALYSIS_YEARS) + 2))

    unique_pmids = list(set(p for year_pmids in all_pmids.values() for p in year_pmids))
    status_text.text(f"Fetching iCite data for {len(unique_pmids):,} unique publications...")

    if not unique_pmids:
        return None, None, None, {}

    icite_data = fetch_icite(unique_pmids, status_fn=lambda msg: status_text.text(msg))
    progress_bar.progress((len(ANALYSIS_YEARS) + 1) / (len(ANALYSIS_YEARS) + 2))

    status_text.text("Computing metrics...")
    metrics_by_year = {}
    for year in ANALYSIS_YEARS:
        year_icite = {pmid: icite_data[pmid] for pmid in all_pmids[year] if pmid in icite_data}
        metrics_by_year[year] = compute_institution_metrics(year_icite)

    combined_icite = {pmid: pub for pmid, pub in icite_data.items()}
    combined_metrics = compute_institution_metrics(combined_icite)
    progress_bar.progress(1.0)

    return metrics_by_year, combined_metrics, icite_data, all_pmids


# ===========================================================================
# Streamlit UI
# ===========================================================================
st.set_page_config(page_title="Institution RCR Analyzer", page_icon="\U0001f3db", layout="wide")

st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Libre+Baskerville:wght@400;700&family=Nunito+Sans:wght@400;500;600;700&display=swap');
    .stApp { background-color: #fafbfc; }
    .main-header { font-family: 'Libre Baskerville', Georgia, serif; font-size: 2.4rem; font-weight: 700; color: #1B4F72; margin-bottom: 0.2rem; letter-spacing: -0.01em; }
    .sub-header { font-family: 'Nunito Sans', sans-serif; font-size: 1.05rem; color: #5D6D7E; margin-bottom: 0.3rem; font-weight: 400; }
    .detail-text { font-family: 'Nunito Sans', sans-serif; font-size: 0.88rem; color: #85929E; margin-bottom: 2rem; line-height: 1.5; }
    .section-label { font-family: 'Nunito Sans', sans-serif; font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: #85929E; margin-bottom: 0.5rem; }
    .stat-card { background: white; border: 1px solid #D5DBDB; border-radius: 12px; padding: 1.25rem 1.5rem; text-align: center; }
    .stat-value { font-family: 'Libre Baskerville', Georgia, serif; font-size: 1.8rem; font-weight: 700; color: #1B4F72; }
    .stat-label { font-family: 'Nunito Sans', sans-serif; font-size: 0.78rem; color: #85929E; text-transform: uppercase; letter-spacing: 0.05em; margin-top: 0.25rem; }
    div[data-testid="stDataFrame"] { border: 1px solid #D5DBDB; border-radius: 8px; }
    button[data-baseweb="tab"] { font-family: 'Nunito Sans', sans-serif !important; font-size: 1rem !important; font-weight: 500 !important; color: #5D6D7E !important; padding: 0.75rem 1.25rem !important; }
    button[data-baseweb="tab"][aria-selected="true"] { color: #1B4F72 !important; font-weight: 700 !important; }
    button[data-baseweb="tab"]:hover { color: #1B4F72 !important; }
    div[data-baseweb="tab-list"] { gap: 0.5rem; }
    .variant-box { background: #EBF5FB; border: 1px solid #AED6F1; border-radius: 8px; padding: 1rem; margin: 0.5rem 0 1rem 0; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">Institution RCR Analyzer</div>', unsafe_allow_html=True)
st.markdown('<div class="sub-header">Analyze publication impact metrics for academic medical centers, hospitals, and schools of medicine.</div>', unsafe_allow_html=True)
st.markdown(
    f'<div class="detail-text">'
    f'This site queries PubMed for all publications listing the institution under any author\'s affiliation '
    f'during the last two full calendar years ({ANALYSIS_YEARS[0]} and {ANALYSIS_YEARS[1]}). '
    f'It then retrieves NIH iCite data to calculate institution-level metrics including the total RCR sum, '
    f'mean RCR per publication, and median RCR per publication.'
    f'</div>',
    unsafe_allow_html=True,
)

# --- Sidebar ---
with st.sidebar:
    st.markdown('<div class="section-label">NCBI Credentials</div>', unsafe_allow_html=True)
    st.caption("Required by NCBI for E-utilities access. Your credentials are not stored.")
    email = st.text_input("Email address", placeholder="you@university.edu")
    api_key = st.text_input("NCBI API key (optional)", type="password", placeholder="Paste your key here")
    if api_key:
        st.success("API key provided \u2014 requests at 10/sec")
    else:
        st.info("No API key \u2014 requests limited to 3/sec")
    st.markdown("---")
    st.markdown('<div class="section-label">How to get an API key</div>', unsafe_allow_html=True)
    st.caption("1. Sign in at [ncbi.nlm.nih.gov](https://www.ncbi.nlm.nih.gov/myncbi/)\n2. Go to Account Settings\n3. Scroll to API Key Management\n4. Click Create API Key")
    st.markdown("---")
    st.markdown('<div class="section-label">About the metrics</div>', unsafe_allow_html=True)
    st.caption(
        "**RCR (Relative Citation Ratio)**: A field-normalized metric where 1.0 = average for the field. "
        "An RCR of 2.0 means the paper is cited twice as much as expected.\n\n"
        "**NIH Percentile**: Where the paper falls among all NIH-funded publications of the same age.\n\n"
        "**Provisional**: Newer papers may have provisional RCR scores that can change as citations accumulate."
    )

# --- Session state ---
if "institutions" not in st.session_state:
    st.session_state.institutions = []
if "pending_variants" not in st.session_state:
    st.session_state.pending_variants = None

# --- Input ---
st.markdown('<div class="section-label">ADD INSTITUTIONS</div>', unsafe_allow_html=True)

st.caption(
    "Enter an institution name and the app will auto-generate common affiliation variants. "
    "You can review, edit, add, or remove variants before adding the institution to your list. "
    "Use \"Add another institution\" to compare multiple institutions."
)

inst_label = st.text_input(
    "Institution name",
    placeholder="e.g. University of California, San Francisco",
    key="inst_label",
)

if st.button("\U0001f50d Generate affiliation variants"):
    if inst_label.strip():
        variants = generate_institution_variants(inst_label.strip())
        st.session_state.pending_variants = {
            "label": inst_label.strip(),
            "variants": variants,
        }
        st.rerun()
    else:
        st.warning("Please enter an institution name.")

# --- Show pending variants for review ---
if st.session_state.pending_variants is not None:
    pending = st.session_state.pending_variants
    st.markdown(f'<div class="section-label">REVIEW VARIANTS FOR: {pending["label"]}</div>', unsafe_allow_html=True)
    st.caption("Check or uncheck variants to include in the search. You can also add custom variants below.")

    # Editable variant list with checkboxes
    selected = []
    for i, v in enumerate(pending["variants"]):
        if st.checkbox(v, value=True, key=f"var_cb_{i}"):
            selected.append(v)

    # Add custom variants
    custom_variants = st.text_input(
        "Add custom variants (comma-separated)",
        placeholder="e.g. My Hospital Name, Another Alias",
        key="custom_variants",
    )

    col_add, col_cancel = st.columns([1, 1])
    with col_add:
        if st.button("\u2705 Add institution with selected variants", use_container_width=True):
            # Merge selected + custom
            all_variants = list(selected)
            if custom_variants:
                for cv in custom_variants.split(","):
                    cv = cv.strip()
                    if cv and cv not in all_variants:
                        all_variants.append(cv)
            if all_variants:
                st.session_state.institutions.append({
                    "label": pending["label"],
                    "variants": all_variants,
                })
                st.session_state.pending_variants = None
                st.rerun()
            else:
                st.warning("Select at least one variant.")
    with col_cancel:
        if st.button("\u274c Cancel", use_container_width=True):
            st.session_state.pending_variants = None
            st.rerun()

# --- Bulk entry ---
st.markdown("")
st.caption("**Bulk entry:** Paste multiple institutions \u2014 one per line. Format: `Label; Variant1, Variant2, ...`")
st.caption("If you only provide a label (no semicolon), variants will be auto-generated.")
bulk_text = st.text_area(
    "Bulk entry (optional)",
    placeholder="University of California, San Francisco\nYale University\nMayo Clinic; Mayo Clinic, Mayo Clinic Rochester, Mayo Foundation",
    height=100, key="bulk_input",
)
if st.button("\u2795 Add all from bulk entry"):
    added = 0
    for line in bulk_text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        if ";" in line:
            parts = line.split(";", 1)
            label = parts[0].strip()
            variants = [v.strip() for v in parts[1].split(",") if v.strip()]
        else:
            label = line
            variants = generate_institution_variants(label)
        if label and variants:
            st.session_state.institutions.append({"label": label, "variants": variants})
            added += 1
    if added:
        st.rerun()

# --- Institution list ---
institutions = st.session_state.institutions

if institutions:
    st.markdown("---")
    st.markdown(f'<div class="section-label">INSTITUTIONS TO ANALYZE \u2014 {len(institutions)} TOTAL</div>', unsafe_allow_html=True)

    for i, inst in enumerate(institutions):
        with st.expander(f"**{inst['label']}** \u2014 {len(inst['variants'])} variants", expanded=False):
            for v in inst["variants"]:
                st.markdown(f"- {v}")

    col_clear, col_run = st.columns([1, 3])
    with col_clear:
        if st.button("\U0001f5d1\ufe0f Clear list", use_container_width=True):
            st.session_state.institutions = []
            st.rerun()
    with col_run:
        run_disabled = not email
        if not email:
            st.warning("Enter your email in the sidebar to run.")
        run_clicked = st.button(
            "\U0001f680 Run Analysis", type="primary",
            use_container_width=True, disabled=run_disabled,
        )

    if run_clicked and email:
        st.markdown("---")
        all_results = []

        for inst_idx, inst in enumerate(institutions):
            st.markdown(f'<div class="section-label">ANALYZING: {inst["label"]} ({len(inst["variants"])} affiliation variants)</div>', unsafe_allow_html=True)
            progress_bar = st.progress(0)
            status_text = st.empty()

            metrics_by_year, combined_metrics, icite_data, pmids_by_year = run_institution_pipeline(
                inst["variants"], email, api_key.strip() if api_key else None,
                progress_bar, status_text,
            )

            if combined_metrics is None:
                st.warning(f"No publications found for {inst['label']}.")
                continue

            status_text.text(f"{inst['label']}: Done!")
            all_results.append({
                "label": inst["label"],
                "variants": inst["variants"],
                "metrics_by_year": metrics_by_year,
                "combined": combined_metrics,
                "icite_data": icite_data,
                "pmids_by_year": pmids_by_year,
            })

        if all_results:
            st.markdown("---")
            st.markdown('<div class="main-header" style="font-size:1.6rem;">Results</div>', unsafe_allow_html=True)
            st.markdown(
                f'<div class="detail-text">Analysis period: {ANALYSIS_YEARS[0]}\u2013{ANALYSIS_YEARS[-1]}</div>',
                unsafe_allow_html=True,
            )

            if len(all_results) > 1:
                st.markdown('<div class="section-label">INSTITUTION COMPARISON</div>', unsafe_allow_html=True)
                comparison = []
                for r in all_results:
                    comparison.append({
                        "Institution": r["label"],
                        "Total Pubs": r["combined"]["total_publications"],
                        "RCR Sum": r["combined"]["rcr_sum"],
                        "Mean RCR": r["combined"]["rcr_mean"],
                        "Median RCR": r["combined"]["rcr_median"],
                        "Mean Percentile": r["combined"]["percentile_mean"],
                        "Total Citations": r["combined"]["citation_total"],
                    })
                st.dataframe(comparison, use_container_width=True, hide_index=True)
                st.markdown("")

            st.markdown("---")
            st.markdown('<div class="section-label">DOWNLOADS</div>', unsafe_allow_html=True)
            cols = st.columns(min(len(all_results), 3))
            for i, r in enumerate(all_results):
                with cols[i % 3]:
                    xlsx_buf = write_institution_xlsx(
                        r["label"], r["icite_data"],
                        r["metrics_by_year"], r["combined"],
                    )
                    safe_name = re.sub(r'[^\w\s-]', '', r["label"]).strip().replace(' ', '_')
                    st.download_button(
                        label=f"\U0001f4e5 {r['label']}",
                        data=xlsx_buf,
                        file_name=f"{safe_name}_RCR_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"dl_inst_{i}",
                    )
